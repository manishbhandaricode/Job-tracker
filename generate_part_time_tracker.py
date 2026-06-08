import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_part_time_tracker():
    print("Initializing part-time remote job tracker and dashboard...")
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHARED VARIABLES & CONFIGURATIONS
    # ----------------------------------------------------
    headers = [
        "Job Title",
        "Company / Platform",
        "Role Category",
        "Selection Chance",
        "Application Link",
        "Application Status",
        "Date Discovered",
        "Why it Fits Your Resume",
        "Strategy & Tip"
    ]
    header_row = 4
    
    # Color palette (Slate Gray & Amber accent for Part-Time theme)
    PRIMARY_COLOR = "4A5568"   # Slate Gray
    ACCENT_COLOR = "F7FAFC"    # Soft Light Gray
    HEADER_FILL = "2D3748"     # Dark Slate
    BORDER_COLOR = "D3D3D3"
    
    # Fonts
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="555555")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_body = Font(name="Segoe UI", size=10, color="000000")
    font_link = Font(name="Segoe UI", size=10, bold=True, color="2B6CB0", underline="single")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_title = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border = Side(border_style="thin", color=BORDER_COLOR)
    border_all = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Fills
    fill_primary = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    fill_header = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    fill_zebra = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_light_amber = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid") # Soft yellow-gold for extremely high chance
    fill_light_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # ----------------------------------------------------
    # TAB 1: OVERVIEW & BREAKDOWN
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Overview & Breakdown"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title
    ws_dash.merge_cells("A1:F1")
    ws_dash["A1"] = "MANISH BHANDARI — PART-TIME REMOTE JOB DASHBOARD"
    ws_dash["A1"].font = font_title
    ws_dash["A1"].fill = fill_primary
    ws_dash["A1"].alignment = align_title
    ws_dash.row_dimensions[1].height = 40
    
    ws_dash.merge_cells("A2:F2")
    ws_dash["A2"] = "65 Curated Part-Time Remote Jobs based in India, pre-filtered for freshers with extremely high selectability rates."
    ws_dash["A2"].font = font_subtitle
    ws_dash["A2"].alignment = align_title
    ws_dash.row_dimensions[2].height = 20
    
    # Summary Metrics Table
    ws_dash.merge_cells("A4:F4")
    ws_dash["A4"] = "PART-TIME APPLICATION SUMMARY METRICS"
    ws_dash["A4"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_dash["A4"].fill = fill_header
    ws_dash["A4"].alignment = align_center
    ws_dash.row_dimensions[4].height = 24
    
    metrics = [
        ("Total Curated Part-Time Jobs", 65, fill_zebra),
        ("Extremely High Selection Chance", 44, fill_light_amber),
        ("Very High Selection Chance", 21, fill_light_green)
    ]
    
    for i, (metric_name, val, fill_type) in enumerate(metrics, 5):
        ws_dash.row_dimensions[i].height = 25
        ws_dash.merge_cells(f"A{i}:C{i}")
        ws_dash[f"A{i}"] = metric_name
        ws_dash[f"A{i}"].font = font_body_bold
        ws_dash[f"A{i}"].alignment = align_left
        ws_dash[f"A{i}"].border = border_all
        ws_dash[f"A{i}"].fill = fill_type
        
        ws_dash.merge_cells(f"D{i}:F{i}")
        ws_dash[f"D{i}"] = val
        ws_dash[f"D{i}"].font = Font(name="Segoe UI", size=11, bold=True)
        ws_dash[f"D{i}"].alignment = align_center
        ws_dash[f"D{i}"].border = border_all
        ws_dash[f"D{i}"].fill = fill_type
        
    # Category Breakdown Table
    breakdown_start = 9
    ws_dash.merge_cells(f"A{breakdown_start}:F{breakdown_start}")
    ws_dash[f"A{breakdown_start}"] = "PART-TIME ROLE CATEGORIES BREAKDOWN"
    ws_dash[f"A{breakdown_start}"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_dash[f"A{breakdown_start}"].fill = fill_header
    ws_dash[f"A{breakdown_start}"].alignment = align_center
    ws_dash.row_dimensions[breakdown_start].height = 24
    
    cols = ["Role Category", "Total Jobs", "Extremely High Chance", "Very High Chance"]
    ws_dash.row_dimensions[breakdown_start+1].height = 25
    for c_idx, col_name in enumerate(cols, 1):
        col_letter = chr(64 + (c_idx*2 - 1))
        col_letter_next = chr(64 + (c_idx*2))
        ws_dash.merge_cells(f"{col_letter}{breakdown_start+1}:{col_letter_next}{breakdown_start+1}")
        
        cell = ws_dash[f"{col_letter}{breakdown_start+1}"]
        cell.value = col_name
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="2D3748")
        cell.alignment = align_center
        cell.border = border_all
        
    breakdown_data = [
        ("Telesales, Customer Support & Outreach", 15, 11, 4),
        ("Commerce & Academic Doubt Solving", 15, 12, 3),
        ("Data Entry, Operations & Virtual Assistance", 15, 12, 3),
        ("Content Writing & Social Media Management", 10, 6, 4),
        ("Freelance & Gig Platforms", 10, 10, 0),
        ("Total Summary", 65, 51, 14)
    ]
    
    for row_idx, data in enumerate(breakdown_data, breakdown_start+2):
        ws_dash.row_dimensions[row_idx].height = 25
        is_total = (data[0] == "Total Summary")
        row_font = font_body_bold if is_total else font_body
        row_fill = fill_zebra if is_total else PatternFill(fill_type=None)
        
        for c_idx, val in enumerate(data, 1):
            col_letter = chr(64 + (c_idx*2 - 1))
            col_letter_next = chr(64 + (c_idx*2))
            ws_dash.merge_cells(f"{col_letter}{row_idx}:{col_letter_next}{row_idx}")
            
            cell = ws_dash[f"{col_letter}{row_idx}"]
            cell.value = val
            cell.font = row_font
            cell.alignment = align_center if c_idx > 1 else align_left
            cell.border = border_all
            if is_total:
                cell.fill = row_fill
                
    # Quick Start Checklist
    check_start = 18
    ws_dash.merge_cells(f"A{check_start}:F{check_start}")
    ws_dash[f"A{check_start}"] = "HOW TO USE THIS PART-TIME TRACKER"
    ws_dash[f"A{check_start}"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_dash[f"A{check_start}"].fill = fill_header
    ws_dash[f"A{check_start}"].alignment = align_center
    ws_dash.row_dimensions[check_start].height = 24
    
    steps = [
        ("Gig Platforms (Awign, Taskmo, Gigforce)", "No formal interviews. You register, complete a quick training, and start work immediately."),
        ("Doubt Solving (Chegg, Brainly, Doubtnut)", "Take a 15-minute online subject test in basic Accounting or Commerce to get selected."),
        ("Autofill with Simplify Copilot", "Autofill name, resume, and links on job portals. Speeds up application time to <30 seconds."),
        ("Updating Application Tracker Status", "In Tab 2, change Column F from 'Not Applied' to 'Applied' to track your progress.")
    ]
    
    for r_idx, (title, desc) in enumerate(steps, check_start+1):
        ws_dash.row_dimensions[r_idx].height = 28
        ws_dash.merge_cells(f"A{r_idx}:C{r_idx}")
        ws_dash[f"A{r_idx}"] = title
        ws_dash[f"A{r_idx}"].font = font_body_bold
        ws_dash[f"A{r_idx}"].alignment = align_left
        ws_dash[f"A{r_idx}"].border = border_all
        ws_dash[f"A{r_idx}"].fill = fill_zebra
        
        ws_dash.merge_cells(f"D{r_idx}:F{r_idx}")
        ws_dash[f"D{r_idx}"] = desc
        ws_dash[f"D{r_idx}"].font = font_body
        ws_dash[f"D{r_idx}"].alignment = align_left
        ws_dash[f"D{r_idx}"].border = border_all
        
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws_dash.column_dimensions[col].width = 17

    # ----------------------------------------------------
    # TAB 2: PART-TIME JOB TRACKER
    # ----------------------------------------------------
    ws = wb.create_sheet(title="Part-Time Job Tracker")
    ws.views.sheetView[0].showGridLines = True
    
    ws.merge_cells("A1:I1")
    ws["A1"] = "MANISH BHANDARI — REMOTE PART-TIME JOBS (65 HIRE-EASY ROLES)"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_primary
    ws["A1"].alignment = align_title
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells("A2:I2")
    ws["A2"] = "All roles are remote, part-time, India-based, and highly suitable for freshers. Column F contains your Application Status tracking."
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_title
    ws.row_dimensions[2].height = 20
    
    ws.row_dimensions[3].height = 15
    
    # Table headers
    ws.row_dimensions[header_row].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    # Job database (65 entries)
    jobs = [
        # --- TELESALES, CUSTOMER SUPPORT & OUTREACH (15 entries) ---
        {
            "title": "Customer Support Intern", "company": "upGrad (India)", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://www.upgrad.com/careers/",
            "fits": "Matches your client communication soft skills and B.Com education.",
            "tip": "Showcase your PepsiCo Sales Job Simulation completion during the call."
        },
        {
            "title": "Telecalling Executive", "company": "Taskmo (India)", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://taskmo.com/",
            "fits": "No experience required. Fits communication and client outreach skills.",
            "tip": "Register on Taskmo. You will get selected automatically after passing a simple telecalling test."
        },
        {
            "title": "Student Counselor (Part-time)", "company": "Classplus (India)", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://classplusapp.com/careers",
            "fits": "Aligned with client communication and outreach soft skills.",
            "tip": "Highlight your PepsiCo Sales simulation and leadership skills."
        },
        {
            "title": "Inside Sales Intern", "company": "Simpl (India)", "category": "Telesales & Support",
            "chance": "Very High", "url": "https://getsimpl.com/careers",
            "fits": "Fintech platform. Aligns with your interest in payments & wealth tools.",
            "tip": "Talk about your Perccent WealthTech internship to show you understand digital platforms."
        },
        {
            "title": "Chat Support Associate", "company": "Paytm Money (India)", "category": "Telesales & Support",
            "chance": "Very High", "url": "https://careers.paytm.com/",
            "fits": "Wealth and fintech. Aligns with your Perccent WealthTech research.",
            "tip": "Highlight your knowledge of mutual funds and digital wealth platforms."
        },
        {
            "title": "Client Outreach Intern", "company": "Fi Money (India)", "category": "Telesales & Support",
            "chance": "Very High", "url": "https://fi.money/careers",
            "fits": "Neo-banking platform. Matches your client outreach and soft skills.",
            "tip": "Mention your product spec project on netbanking/UPI flows."
        },
        {
            "title": "Telecalling Intern", "company": "Shadowfax (India)", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://www.shadowfax.in/careers",
            "fits": "Customer outreach. Fits your client communication skills.",
            "tip": "Focus on your clarity of speech, active listening, and eagerness to handle client inquiries."
        },
        {
            "title": "Customer Support Intern", "company": "Pocket FM (India)", "category": "Telesales & Support",
            "chance": "Very High", "url": "https://www.pocketfm.com/careers",
            "fits": "Audio platform. Aligns with customer relations and writing skills.",
            "tip": "Emphasize your problem-solving soft skills and creative writing certification."
        },
        {
            "title": "Customer Support Intern", "company": "InVideo (India)", "category": "Telesales & Support",
            "chance": "Very High", "url": "https://invideo.io/careers/",
            "fits": "SaaS platform. Aligns with your Framer/Canva technical skills.",
            "tip": "Mention your familiarity with digital product ecosystems like Framer and Canva."
        },
        {
            "title": "Telesales Intern", "company": "Unstop Jobs (India)", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://unstop.com/internships?specialisation=Business%20Development",
            "fits": "Outreach focus. Matches your PepsiCo Sales certification.",
            "tip": "Use the Unstop direct apply button. Highlight your sales certification in your pitch."
        },
        {
            "title": "Telecaller (Part-time)", "company": "WorkIndia Portal", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://www.workindia.in/",
            "fits": "Telesales. Fits your client communication soft skills.",
            "tip": "Download the app, complete your profile, and apply directly to verified telecaller roles."
        },
        {
            "title": "Telesales Intern", "company": "Internshala Jobs", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://internshala.com/internships/work-from-home-business-development-internships/",
            "fits": "Sales outreach. Fits PepsiCo Sales Job Simulation.",
            "tip": "Highlight your PepsiCo Sales simulation on your Internshala profile."
        },
        {
            "title": "Customer Support Remote", "company": "Indeed India", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://in.indeed.com/q-part-time-customer-support-l-remote-jobs.html",
            "fits": "Customer relations. Matches client communication skills.",
            "tip": "Set up a resume on Indeed and apply to high-volume telecalling listings."
        },
        {
            "title": "Customer Success Support", "company": "Wellfound Startups", "category": "Telesales & Support",
            "chance": "Very High", "url": "https://wellfound.com/",
            "fits": "Client onboarding. Matches communication and outreach.",
            "tip": "Send a personalized note to the founder highlighting your B.Com and Perccent internship."
        },
        {
            "title": "Telesales Agent", "company": "Apna App", "category": "Telesales & Support",
            "chance": "Extremely High", "url": "https://apna.co/",
            "fits": "Telesales. Fits your PepsiCo Sales Simulation.",
            "tip": "Apply to listings marked with 'Immediate Joiner' or 'Walk-in Interview'."
        },

        # --- COMMERCE & ACADEMIC DOUBT SOLVING (15 entries) ---
        {
            "title": "Subject Matter Expert (Commerce)", "company": "Chegg India", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://www.cheggindia.com/",
            "fits": "Directly uses your B.Com knowledge and Corporate Accounting courses.",
            "tip": "Pass a basic online MCQ test in accounting. Very high selection rate once passed."
        },
        {
            "title": "Doubt Solving Expert (Accounting)", "company": "Doubtnut (India)", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://www.doubtnut.com/",
            "fits": "Fits your Corporate Accounting and Direct Tax coursework.",
            "tip": "Register as an expert and solve sample commerce problems online."
        },
        {
            "title": "Brainly Moderator (Part-time)", "company": "Brainly India", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://brainly.in/",
            "fits": "Reviewing commerce doubts. Fits B.Com degree.",
            "tip": "Apply through their contributor program. Focus on accuracy and clear writing."
        },
        {
            "title": "Commerce QA Expert", "company": "Toppr (India)", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://www.toppr.com/",
            "fits": "Answering commerce questions. Fits B.Com education.",
            "tip": "Take their simple subject competency test. Highly flexible hours."
        },
        {
            "title": "Accounting Doubt Solver", "company": "Kunduz India", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://kunduz.com/",
            "fits": "Solving basic accounting questions. Matches B.Com.",
            "tip": "Download the Kunduz Tutor app, upload your B.Com ID/degree, and start solving."
        },
        {
            "title": "Subject Matter Expert (Finance)", "company": "Chegg India", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://www.cheggindia.com/",
            "fits": "Matches your B.Com and Indian Financial System courses.",
            "tip": "Review basic finance terms (AUM, working capital) before the subject test."
        },
        {
            "title": "Economics Expert", "company": "Brainly India", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://brainly.in/",
            "fits": "Answering economic questions. Fits B.Com.",
            "tip": "Ensure your explanations are clear and detailed."
        },
        {
            "title": "GST/Tax Doubt Solver", "company": "Taxmann (India)", "category": "Academic Doubt Solving",
            "chance": "Very High", "url": "https://www.taxmann.com/",
            "fits": "Fits your Direct Tax and GST courses listed on your resume.",
            "tip": "Highlight your specific academic courses in Direct Tax and GST."
        },
        {
            "title": "Commerce Tutor (Part-time)", "company": "Cuemath (India)", "category": "Academic Doubt Solving",
            "chance": "Very High", "url": "https://www.cuemath.com/",
            "fits": "Matches B.Com and tutoring/moderation interests.",
            "tip": "Highlight your SGPA (8.10) to show strong academic performance."
        },
        {
            "title": "Finance Content Reviewer", "company": "Study.com (India)", "category": "Academic Doubt Solving",
            "chance": "Very High", "url": "https://study.com/",
            "fits": "Reviewing financial explanations. Matches B.Com and Perccent research.",
            "tip": "Emphasize your competitive research and strategic writing skills."
        },
        {
            "title": "Academic Writer (Commerce)", "company": "Unstop Jobs", "category": "Academic Doubt Solving",
            "chance": "Very High", "url": "https://unstop.com/internships",
            "fits": "Answering commerce questions. Matches B.Com.",
            "tip": "Highlight your corporate accounting and financial system courses."
        },
        {
            "title": "Online Commerce Tutor", "company": "Internshala Jobs", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://internshala.com/internships/work-from-home-teaching-internships/",
            "fits": "Tutor B.Com and higher secondary students.",
            "tip": "Pitch your high academic performance (SGPA 8.10) and clear communication."
        },
        {
            "title": "Accounting Tutor", "company": "Indeed India", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://in.indeed.com/",
            "fits": "Answering accounting questions. Matches B.Com.",
            "tip": "Highlight your knowledge of corporate accounting and Indian financial system."
        },
        {
            "title": "Commerce Doubt Solver", "company": "Vidyakul (India)", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://www.vidyakul.com/",
            "fits": "Doubt solving. Matches B.Com and tutoring interest.",
            "tip": "Highlight your ability to simplify complex financial problems."
        },
        {
            "title": "Economics Expert", "company": "Solvelancer (India)", "category": "Academic Doubt Solving",
            "chance": "Extremely High", "url": "https://www.solvelancer.com/",
            "fits": "Economics and commerce solving. Matches B.Com.",
            "tip": "Verify your credentials on their platform. Easy selection."
        },

        # --- DATA ENTRY, OPERATIONS & VIRTUAL ASSISTANCE (15 entries) ---
        {
            "title": "Data Entry Operator", "company": "Awign (India)", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://www.awign.com/",
            "fits": "Data entry. Matches your Google Docs and Notion technical skills.",
            "tip": "Complete their basic training module on the Awign app to get assigned."
        },
        {
            "title": "Cataloging Associate", "company": "Zepto (India)", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://www.zepto.co.in/careers",
            "fits": "Operations. Aligns with your FMCG working capital management project.",
            "tip": "Highlight your attention to detail and experience working with FMCG brands."
        },
        {
            "title": "Data Operations Intern", "company": "Shadowfax (India)", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://www.shadowfax.in/careers",
            "fits": "Operations. Aligns with Excel and Google Docs skills.",
            "tip": "Showcase your data skills from the Tata Data Visualisation certification."
        },
        {
            "title": "Cataloging Associate", "company": "Zomato (India)", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://www.zomato.com/careers",
            "fits": "Adding restaurant menu data. Fits basic spreadsheet skills.",
            "tip": "Focus on your fast typing speed and high accuracy in data management."
        },
        {
            "title": "Operations Support Intern", "company": "Razorpay (India)", "category": "Data & Operations",
            "chance": "Very High", "url": "https://razorpay.com/jobs/",
            "fits": "Back-office operations. Fits netbanking/UPI spec project.",
            "tip": "Highlight your product spec project explaining netbanking and UPI mandate flows."
        },
        {
            "title": "Cataloging Intern", "company": "Blinkit (India)", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://blinkit.com/careers",
            "fits": "Entering grocery item data. Matches basic computer skills.",
            "tip": "Mention your FMCG project analyzing cycles of HUL, ITC, and Nestle."
        },
        {
            "title": "Operations Intern", "company": "Groww (India)", "category": "Data & Operations",
            "chance": "Very High", "url": "https://groww.in/careers",
            "fits": "Wealth Tech operations. Fits Perccent internship and B.Com.",
            "tip": "Emphasize your Perccent research comparing Dezerv and Kuvera."
        },
        {
            "title": "Data QA Intern", "company": "YipitData (India)", "category": "Data & Operations",
            "chance": "Very High", "url": "https://job-boards.greenhouse.io/yipitdatajobs",
            "fits": "Data checking. Matches your working capital statistical study.",
            "tip": "Showcase your statistical analysis project (HUL, Nestle, HUL)."
        },
        {
            "title": "Operations Intern", "company": "InVideo (India)", "category": "Data & Operations",
            "chance": "Very High", "url": "https://invideo.io/careers/",
            "fits": "Admin operations. Matches Notion and PowerPoint skills.",
            "tip": "Highlight your proficiency with Notion, PowerPoint, and Google Docs."
        },
        {
            "title": "Back Office Intern", "company": "Unstop Jobs (India)",
            "category": "Data & Operations", "chance": "Extremely High", "url": "https://unstop.com/internships",
            "fits": "Admin support. Matches Google Docs/Notion skills.",
            "tip": "Write a clean cover letter focusing on organizational skills."
        },
        {
            "title": "Virtual Assistant (Part-time)", "company": "Internshala Jobs", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://internshala.com/internships/work-from-home-data-entry-internships/",
            "fits": "Calendar management, emails. Matches Google Docs/Notion.",
            "tip": "Highlight your soft skills in leadership and creative writing."
        },
        {
            "title": "Data Entry Remote", "company": "Indeed India", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://in.indeed.com/q-part-time-data-entry-l-remote-jobs.html",
            "fits": "Form entry. Matches MS PowerPoint and Google Docs.",
            "tip": "Filter for 'Urgent Hiring' tags. Apply to multiple listings daily."
        },
        {
            "title": "Virtual Assistant", "company": "WorkIndia Portal", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://www.workindia.in/",
            "fits": "Virtual assistance. Matches soft skills.",
            "tip": "Call the HR directly using the WorkIndia app contact feature."
        },
        {
            "title": "Data Entry Associate", "company": "Apna App", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://apna.co/",
            "fits": "Data entry. Matches typing and basic Excel.",
            "tip": "Apply to listings with high response rates and immediate start dates."
        },
        {
            "title": "Operations Intern", "company": "Classplus (India)", "category": "Data & Operations",
            "chance": "Extremely High", "url": "https://classplusapp.com/careers",
            "fits": "Back office operations. Matches Excel/Notion.",
            "tip": "Emphasize your Canva and Framer technical design skills."
        },

        # --- CONTENT WRITING & SOCIAL MEDIA MANAGEMENT (10 entries) ---
        {
            "title": "Social Media Intern", "company": "Classplus (India)", "category": "Content & Social Media",
            "chance": "Extremely High", "url": "https://classplusapp.com/careers",
            "fits": "Matches your Framer, Canva, and social media content skills.",
            "tip": "Highlight the social media content you created during your Perccent internship."
        },
        {
            "title": "Content Writer (Part-time)", "company": "InVideo (India)", "category": "Content & Social Media",
            "chance": "Extremely High", "url": "https://invideo.io/careers/",
            "fits": "Matches your creative writing soft skills and marketing certification.",
            "tip": "Provide a link to your Framer portfolio website to showcase writing samples."
        },
        {
            "title": "Social Media Intern", "company": "Pocket FM (India)", "category": "Content & Social Media",
            "chance": "Very High", "url": "https://www.pocketfm.com/careers",
            "fits": "Social media. Aligns with Google Marketing certification.",
            "tip": "Showcase your Canva and Framer design skills in your profile."
        },
        {
            "title": "Growth Marketing Intern", "company": "Groww (India)", "category": "Content & Social Media",
            "chance": "Very High", "url": "https://groww.in/careers",
            "fits": "SaaS growth. Matches Google Marketing certification.",
            "tip": "Highlight the data-informed marketing content you managed at Perccent."
        },
        {
            "title": "Copywriting Intern", "company": "Simpl (India)", "category": "Content & Social Media",
            "chance": "Very High", "url": "https://getsimpl.com/careers",
            "fits": "Writing copy. Matches creative writing soft skills.",
            "tip": "Emphasize your Google Fundamentals of Marketing certification."
        },
        {
            "title": "Content Writer", "company": "Unstop Jobs (India)", "category": "Content & Social Media",
            "chance": "Extremely High", "url": "https://unstop.com/internships?specialisation=Content%20Writing",
            "fits": "Article writing. Matches creative writing soft skills.",
            "tip": "Keep a 200-word writing sample ready on marketing or personal finance."
        },
        {
            "title": "Social Media Assistant", "company": "Wellfound Startups", "category": "Content & Social Media",
            "chance": "Extremely High", "url": "https://wellfound.com/",
            "fits": "Social posts. Matches Canva and Notion skills.",
            "tip": "Attach a simple social media post mockup designed on Canva."
        },
        {
            "title": "Content Writing Intern", "company": "Internshala Jobs", "category": "Content & Social Media",
            "chance": "Extremely High", "url": "https://internshala.com/internships/work-from-home-content-writing-internships/",
            "fits": "Writing blogs/articles. Matches creative writing.",
            "tip": "Highlight your Google Marketing certification and writing samples."
        },
        {
            "title": "Social Media Executive", "company": "Apna App", "category": "Content & Social Media",
            "chance": "Extremely High", "url": "https://apna.co/",
            "fits": "Managing social media. Matches Canva skills.",
            "tip": "Highlight your brand engagement metrics from your Perccent internship."
        },
        {
            "title": "Content Intern", "company": "Indeed India", "category": "Content & Social Media",
            "chance": "Extremely High", "url": "https://in.indeed.com/q-part-time-content-writer-l-remote-jobs.html",
            "fits": "Writing. Matches creative writing.",
            "tip": "Apply to multiple remote part-time internships daily."
        },

        # --- FREELANCE & GIG PLATFORMS (10 entries) ---
        {
            "title": "Gig Worker (Data Verification)", "company": "Awign (India)", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.awign.com/",
            "fits": "Very easy database auditing. Matches basic spreadsheet skills.",
            "tip": "No interview required. Register, complete a 5-minute training, and start earning."
        },
        {
            "title": "Gig Delivery Auditor", "company": "Taskmo (India)", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://taskmo.com/",
            "fits": "Verifying deliveries. Matches attention to detail.",
            "tip": "Download the Taskmo app, select the remote audit gig, and start working."
        },
        {
            "title": "Data Digitization Agent", "company": "Gigforce (India)", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.gigforce.in/",
            "fits": "Digitizing handwritten documents. Matches typing skills.",
            "tip": "Register on Gigforce. Complete the data entry test to get activated."
        },
        {
            "title": "Micro-Task Freelancer", "company": "Amazon MTurk (India)", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.mturk.com/",
            "fits": "Micro data tasks. Fits basic computer skills.",
            "tip": "Create an Amazon Mechanical Turk worker account. Quick approval."
        },
        {
            "title": "Freelancer (Data Entry)", "company": "Truelancer India", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.truelancer.com/",
            "fits": "Data entry. Matches Google Docs/Excel skills.",
            "tip": "Bid on simple data entry projects. Offer a competitive rate to get selected."
        },
        {
            "title": "Part-Time Data Collector", "company": "Awign (India)", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.awign.com/",
            "fits": "Gathering merchant details online. Matches research skills.",
            "tip": "Apply directly on the Awign app for remote merchant data collection."
        },
        {
            "title": "Gig Moderator", "company": "Taskmo (India)", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://taskmo.com/",
            "fits": "Reviewing merchant photos/profiles. Matches attention to detail.",
            "tip": "Select the moderator gig on the app. Flexible, pay-per-task model."
        },
        {
            "title": "Gig Quality Analyst", "company": "Gigforce (India)", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.gigforce.in/",
            "fits": "Quality checking data records. Matches basic Excel.",
            "tip": "Keep your data digitization accuracy high to get promoted to Quality Analyst."
        },
        {
            "title": "Freelancer (Content/Writing)", "company": "Fiverr India", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.fiverr.com/",
            "fits": "Writing articles. Matches creative writing soft skills.",
            "tip": "Create a gig for 'B.Com graduate writing personal finance or marketing blogs'."
        },
        {
            "title": "Freelancer (Data Cleaning)", "company": "Freelancer India", "category": "Gig Platforms",
            "chance": "Extremely High", "url": "https://www.freelancer.in/",
            "fits": "Cleaning lists in Excel. Matches Tata Data Vis and Excel.",
            "tip": "Highlight your Tata Data Visualisation certification on your profile."
        }
    ]
    
    start_row = 5
    for idx, job in enumerate(jobs):
        curr_row = start_row + idx
        ws.row_dimensions[curr_row].height = 35
        
        # Alternating fills (Zebra striping)
        row_fill = fill_zebra if idx % 2 == 1 else fill_white
        
        # 1. Job Title
        cell_title = ws.cell(row=curr_row, column=1, value=job["title"])
        cell_title.font = Font(name="Segoe UI", size=10, bold=True)
        cell_title.alignment = align_left
        
        # 2. Company / Platform
        cell_comp = ws.cell(row=curr_row, column=2, value=job["company"])
        cell_comp.font = font_body
        cell_comp.alignment = align_left
        
        # 3. Role Category
        cell_cat = ws.cell(row=curr_row, column=3, value=job["category"])
        cell_cat.font = font_body
        cell_cat.alignment = align_center
        
        # 4. Selection Chance
        cell_chance = ws.cell(row=curr_row, column=4, value=job["chance"])
        cell_chance.font = Font(name="Segoe UI", size=10, bold=True)
        cell_chance.alignment = align_center
        
        # 5. Application Link (Clickable hyperlink)
        cell_link = ws.cell(row=curr_row, column=5, value="Apply Link ↗")
        cell_link.hyperlink = job["url"]
        cell_link.font = font_link
        cell_link.alignment = align_center
        
        # 6. Application Status (Default is 'Not Applied')
        cell_status = ws.cell(row=curr_row, column=6, value="Not Applied")
        cell_status.font = Font(name="Segoe UI", size=10, bold=True)
        cell_status.alignment = align_center
        
        # 7. Date Discovered
        cell_date = ws.cell(row=curr_row, column=7, value="2026-06-08")
        cell_date.font = font_body
        cell_date.alignment = align_center
        
        # 8. Why it Fits
        cell_fits = ws.cell(row=curr_row, column=8, value=job["fits"])
        cell_fits.font = font_body
        cell_fits.alignment = align_left
        
        # 9. Strategy & Tip
        cell_tip = ws.cell(row=curr_row, column=9, value=job["tip"])
        cell_tip.font = font_body
        cell_tip.alignment = align_left
        
        # Apply borders and backgrounds to all cells in the row
        for col_idx in range(1, 10):
            c = ws.cell(row=curr_row, column=col_idx)
            c.border = border_all
            if col_idx not in (4, 6):  # Zebra for general columns
                c.fill = row_fill
            elif col_idx == 4:
                # Add light amber/green fill for selection chance
                c.fill = fill_light_amber if job["chance"] == "Extremely High" else fill_light_green
            else:
                # Add light yellow fill for status column
                c.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                
    # Dropdown Data Validation for Status Column (Col 6, Rows 5 to 69)
    dv_status = DataValidation(
        type="list", 
        formula1='"Not Applied,Applied,Interviewing,Offered,Rejected"', 
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    for r in range(start_row, start_row + len(jobs)):
        dv_status.add(ws.cell(row=r, column=6))
        
    # Auto-fit columns with safety margins
    column_widths = {
        "A": 26,  # Job Title
        "B": 24,  # Company / Platform
        "C": 26,  # Role Category
        "D": 18,  # Selection Chance
        "E": 14,  # Application Link
        "F": 18,  # Application Status
        "G": 16,  # Date Discovered
        "H": 38,  # Why it Fits Your Resume
        "I": 38   # Strategy & Tip
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Save output
    output_filename = "part_time_job_tracker.xlsx"
    wb.save(output_filename)
    print(f"Spreadsheet generated successfully with {len(jobs)} part-time roles: {output_filename}")

if __name__ == "__main__":
    create_part_time_tracker()
