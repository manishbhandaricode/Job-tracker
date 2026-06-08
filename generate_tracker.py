import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_tracker():
    print("Initializing multi-sheet job tracker and dashboard...")
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHARED VARIABLES & CONFIGURATIONS
    # ----------------------------------------------------
    headers = [
        "Job Title",
        "Company / Platform (India)",
        "Role Category",
        "Employment Type",
        "Application Link",
        "Activity Status",
        "Date Discovered",
        "Why it Fits Your Resume",
        "Application Strategy & Tip"
    ]
    header_row = 4
    
    # Color palette
    PRIMARY_COLOR = "2B4C7E"   # Deep Steel Blue
    ACCENT_COLOR = "EAEFF5"    # Soft Blue-Gray
    HEADER_FILL = "1F4E78"     # Navy Blue
    WHITE_FILL = "FFFFFF"
    BORDER_COLOR = "D3D3D3"
    
    # Fonts
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="555555")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_body = Font(name="Segoe UI", size=10, color="000000")
    font_link = Font(name="Segoe UI", size=10, bold=True, color="1F4E78", underline="single")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_title = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border = Side(border_style="thin", color=BORDER_COLOR)
    border_all = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Fills
    fill_primary = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    fill_header = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    fill_zebra = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_light_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_light_yellow = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    
    # ----------------------------------------------------
    # TAB 1: DASHBOARD & BREAKDOWN
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Dashboard & Breakdown"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # --- 1. Dashboard Title ---
    ws_dash.merge_cells("A1:F1")
    ws_dash["A1"] = "MANISH BHANDARI — REMOTE JOB APPLICATION DASHBOARD"
    ws_dash["A1"].font = font_title
    ws_dash["A1"].fill = fill_primary
    ws_dash["A1"].alignment = align_title
    ws_dash.row_dimensions[1].height = 40
    
    ws_dash.merge_cells("A2:F2")
    ws_dash["A2"] = "Summary & category breakdown of 56 curated remote fresher jobs matching your Commerce & WealthTech profile."
    ws_dash["A2"].font = font_subtitle
    ws_dash["A2"].alignment = align_title
    ws_dash.row_dimensions[2].height = 20
    
    # --- 2. Metrics Cards ---
    metrics = [
        ("Total Curated Jobs", 56, fill_zebra),
        ("Full-Time Positions", 40, fill_light_green),
        ("Part-Time / Internships", 16, fill_light_yellow)
    ]
    
    ws_dash.merge_cells("A4:F4")
    ws_dash["A4"] = "APPLICATION SUMMARY METRICS"
    ws_dash["A4"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_dash["A4"].fill = fill_header
    ws_dash["A4"].alignment = align_center
    ws_dash.row_dimensions[4].height = 24
    
    for i, (metric_name, val, fill_type) in enumerate(metrics, 5):
        ws_dash.row_dimensions[i].height = 25
        ws_dash.merge_cells(f"A{i}:C{i}")
        ws_dash[f"A{i}"] = metric_name
        ws_dash[f"A{i}"].font = font_body_bold
        ws_dash[f"A{i}"].alignment = align_left
        ws_dash[f"A{i}"].border = border_all
        ws_dash[f"A{i}"].fill = fill_type
        
        ws_dash.merge_cells(f"D{i}:F{i}")
        ws_dash[f"D{i}"] = val
        ws_dash[f"D{i}"].font = Font(name="Segoe UI", size=11, bold=True)
        ws_dash[f"D{i}"].alignment = align_center
        ws_dash[f"D{i}"].border = border_all
        ws_dash[f"D{i}"].fill = fill_type
        
    # --- 3. Category Breakdown Table ---
    breakdown_start = 9
    ws_dash.merge_cells(f"A{breakdown_start}:F{breakdown_start}")
    ws_dash[f"A{breakdown_start}"] = "ROLE CATEGORY & COMMITMENT BREAKDOWN"
    ws_dash[f"A{breakdown_start}"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_dash[f"A{breakdown_start}"].fill = fill_header
    ws_dash[f"A{breakdown_start}"].alignment = align_center
    ws_dash.row_dimensions[breakdown_start].height = 24
    
    cols = ["Role Category", "Total Jobs", "Full-Time (FT)", "Part-Time (PT)"]
    ws_dash.row_dimensions[breakdown_start+1].height = 25
    for c_idx, col_name in enumerate(cols, 1):
        # Merge for spacing
        col_letter = chr(64 + (c_idx*2 - 1))  # A, C, E
        col_letter_next = chr(64 + (c_idx*2)) # B, D, F
        range_str = f"{col_letter}{breakdown_start+1}:{col_letter_next}{breakdown_start+1}"
        ws_dash.merge_cells(range_str)
        
        cell = ws_dash[f"{col_letter}{breakdown_start+1}"]
        cell.value = col_name
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")
        cell.alignment = align_center
        cell.border = border_all
        
    breakdown_data = [
        ("Product Management", 12, 8, 4),
        ("Business Analysis", 12, 9, 3),
        ("Finance & Operations", 12, 10, 2),
        ("Marketing & Growth", 10, 6, 4),
        ("Sales & Business Development", 10, 7, 3),
        ("Total Summary", 56, 40, 16)
    ]
    
    for row_idx, data in enumerate(breakdown_data, breakdown_start+2):
        ws_dash.row_dimensions[row_idx].height = 25
        is_total = (data[0] == "Total Summary")
        row_font = font_body_bold if is_total else font_body
        row_fill = fill_zebra if is_total else PatternFill(fill_type=None)
        
        for c_idx, val in enumerate(data, 1):
            col_letter = chr(64 + (c_idx*2 - 1))
            col_letter_next = chr(64 + (c_idx*2))
            ws_dash.merge_cells(f"{col_letter}{row_idx}:{col_letter_next}{row_idx}")
            
            cell = ws_dash[f"{col_letter}{row_idx}"]
            cell.value = val
            cell.font = row_font
            cell.alignment = align_center if c_idx > 1 else align_left
            cell.border = border_all
            if is_total:
                cell.fill = row_fill
                
    # --- 4. Resume Alignment Guide ---
    guide_start = 17
    ws_dash.merge_cells(f"A{guide_start}:F{guide_start}")
    ws_dash[f"A{guide_start}"] = "HOW THESE ROLES ALIGN WITH YOUR RESUME"
    ws_dash[f"A{guide_start}"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_dash[f"A{guide_start}"].fill = fill_header
    ws_dash[f"A{guide_start}"].alignment = align_center
    ws_dash.row_dimensions[guide_start].height = 24
    
    alignments = [
        ("Product Management", "Fits your Core Wealth Tools Product Spec project (Step-Up SIP, NetBanking/UPI flow acceptance criteria)."),
        ("Business Analysis", "Leverages your Perccent competitor research (Dezerv, Kuvera fee models) and Tata Data Vis certification."),
        ("Finance & Operations", "Built for your B.Com degree and FMCG Working Capital project (analyzing HUL, ITC, Nestle)."),
        ("Marketing & Growth", "Aligned with your Google Fundamentals of Marketing certification and Perccent LinkedIn/IG content work."),
        ("Sales & Business Development", "Leverages your PepsiCo Sales Simulation certification and client outreach soft skills.")
    ]
    
    for r_idx, (cat, desc) in enumerate(alignments, guide_start+1):
        ws_dash.row_dimensions[r_idx].height = 30
        # Category Column (A-B)
        ws_dash.merge_cells(f"A{r_idx}:B{r_idx}")
        ws_dash[f"A{r_idx}"] = cat
        ws_dash[f"A{r_idx}"].font = font_body_bold
        ws_dash[f"A{r_idx}"].alignment = align_left
        ws_dash[f"A{r_idx}"].border = border_all
        ws_dash[f"A{r_idx}"].fill = fill_zebra
        
        # Description Column (C-F)
        ws_dash.merge_cells(f"C{r_idx}:F{r_idx}")
        ws_dash[f"C{r_idx}"] = desc
        ws_dash[f"C{r_idx}"].font = font_body
        ws_dash[f"C{r_idx}"].alignment = align_left
        ws_dash[f"C{r_idx}"].border = border_all
        
    # --- 5. Semi-Autofill Checklist ---
    check_start = 24
    ws_dash.merge_cells(f"A{check_start}:F{check_start}")
    ws_dash[f"A{check_start}"] = "QUICK START: SEMI-AUTOMATED APPLICATION WORKFLOW"
    ws_dash[f"A{check_start}"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws_dash[f"A{check_start}"].fill = fill_primary
    ws_dash[f"A{check_start}"].alignment = align_center
    ws_dash.row_dimensions[check_start].height = 24
    
    steps = [
        ("Step 1: Install", "Install the free 'Simplify Copilot' extension in your Chrome browser."),
        ("Step 2: Profile", "Create your profile, upload your resume, and save common details."),
        ("Step 3: Autofill", "Click 'Apply Link' in Tab 2, then click 'Autofill Application' on the job page."),
        ("Step 4: Update", "Change 'Activity Status' in Tab 2 from 'Pending' to 'Applied' after submitting.")
    ]
    
    for r_idx, (step_title, step_desc) in enumerate(steps, check_start+1):
        ws_dash.row_dimensions[r_idx].height = 28
        ws_dash.merge_cells(f"A{r_idx}:B{r_idx}")
        ws_dash[f"A{r_idx}"] = step_title
        ws_dash[f"A{r_idx}"].font = font_body_bold
        ws_dash[f"A{r_idx}"].alignment = align_center
        ws_dash[f"A{r_idx}"].border = border_all
        ws_dash[f"A{r_idx}"].fill = fill_zebra
        
        ws_dash.merge_cells(f"C{r_idx}:F{r_idx}")
        ws_dash[f"C{r_idx}"] = step_desc
        ws_dash[f"C{r_idx}"].font = font_body
        ws_dash[f"C{r_idx}"].alignment = align_left
        ws_dash[f"C{r_idx}"].border = border_all
        
    # Adjust column widths for Dashboard Tab
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws_dash.column_dimensions[col].width = 17
 
    # ----------------------------------------------------
    # TAB 2: JOB TRACKER
    # ----------------------------------------------------
    ws = wb.create_sheet(title="Job Tracker")
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Add Title Banner for Tracker tab
    ws.merge_cells("A1:I1")
    ws["A1"] = "MANISH BHANDARI — INDIA REMOTE JOB TRACKER (56 CURATED FT & PT ROLES)"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_primary
    ws["A1"].alignment = align_title
    ws.row_dimensions[1].height = 40
    
    # Add Subtitle for Tracker tab
    ws.merge_cells("A2:I2")
    ws["A2"] = "Double-click cells in Column D to select Employment Type and Column F to update Application Status. Click Apply Link to open portal."
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_title
    ws.row_dimensions[2].height = 20
    
    # Row 3 is blank spacer
    ws.row_dimensions[3].height = 15
    
    # Column headers for Tracker tab
    ws.row_dimensions[header_row].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    # Expanded Job Data List (56 entries)
    jobs = [
        # --- PRODUCT MANAGEMENT & PRODUCT ANALYSIS (12 entries) ---
        {
            "title": "Associate Product Manager",
            "company": "Groww (India)",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://groww.in/careers",
            "fits": "WealthTech platform. Aligns with your Perccent internship and SIP spec project.",
            "tip": "Attach your 'Product Specification: Core Wealth Tools' project and highlight acceptance criteria."
        },
        {
            "title": "Product Analyst",
            "company": "Kayzen (India)",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/kayzen",
            "fits": "Focus on data analytics and SQL. Leverages your Tata Data Visualisation certification.",
            "tip": "Highlight your experience authoring product specs and translating strategic research into dashboards."
        },
        {
            "title": "Product Analyst",
            "company": "Kernel DAO (India)",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/kerneldao",
            "fits": "DeFi & liquidity focus. Aligns with your CRED x Kuvera M&A project and Perccent WealthTech experience.",
            "tip": "Discuss your analysis of SEBI RIA capabilities and monetization challenges in CRED/Kuvera project."
        },
        {
            "title": "Associate Product Manager",
            "company": "YipitData (India)",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://job-boards.greenhouse.io/yipitdatajobs",
            "fits": "Market intelligence. Aligns with your competitive benchmarking skills.",
            "tip": "Highlight your WealthTech competitor analysis and data-driven approach."
        },
        {
            "title": "Product Analyst Intern",
            "company": "Simpl (India)",
            "category": "Product Management",
            "employment": "Part-time",
            "url": "https://getsimpl.com/careers",
            "fits": "BNPL/FinTech. Fits your understanding of digital wealth and payments. Flexible internship.",
            "tip": "Highlight NetBanking and UPI mandate flow designs from your wealth tools project."
        },
        {
            "title": "Product Analyst",
            "company": "Pearl (India)",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/pearl",
            "fits": "Global AI network. Matches your data-driven problem-solving projects.",
            "tip": "Highlight your working capital statistical correlation study to prove analytical depth."
        },
        {
            "title": "APM Intern (Remote)",
            "company": "Acko (India)",
            "category": "Product Management",
            "employment": "Part-time",
            "url": "https://www.acko.com/careers/",
            "fits": "Digital insurance startup. Fits your product spec writing and analytics skills.",
            "tip": "Focus on your structured problem-solving and documentation skills (PRDs)."
        },
        {
            "title": "Product Management Intern",
            "company": "Unstop Jobs (India)",
            "category": "Product Management",
            "employment": "Part-time",
            "url": "https://unstop.com/internships?specialisation=Product%20Management",
            "fits": "Curated remote product internships for freshers with flexible/part-time hours.",
            "tip": "Make sure your Framer project links are clickable on your resume."
        },
        {
            "title": "Junior Product Analyst",
            "company": "Fi Money (India)",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://fi.money/careers",
            "fits": "Neo-banking platform. Matches your interest in digital wealth and UPI mandates.",
            "tip": "Showcase your 'Product Specification' project and your Perccent internship."
        },
        {
            "title": "Associate Product Manager",
            "company": "Jupiter Money (India)",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://jupiter.money/careers/",
            "fits": "Digital banking & wealth management. Perfect fit for B.Com with product specs.",
            "tip": "Highlight your understanding of Step-Up SIP and CAGR toggle logic."
        },
        {
            "title": "Product Intern",
            "company": "NextLeap (India)",
            "category": "Product Management",
            "employment": "Part-time",
            "url": "https://nextleap.app/",
            "fits": "Product-focused community. Great for freshers, flexible hours.",
            "tip": "Highlight acceptance criteria writing experience and structured document writing."
        },
        {
            "title": "APM Remote India",
            "company": "Indeed India",
            "category": "Product Management",
            "employment": "Full-time",
            "url": "https://in.indeed.com/q-associate-product-manager-l-remote-jobs.html",
            "fits": "Aggregated active remote APM roles in India.",
            "tip": "Set daily alerts and apply within 24 hours of posting for maximum visibility."
        },

        # --- BUSINESS ANALYSIS & DATA/MARKET RESEARCH (12 entries) ---
        {
            "title": "Business Analyst (Remote)",
            "company": "WIN Home Inspection (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/winhomeinspection",
            "fits": "Focus on research and strategic initiatives. Fits your B.Com.",
            "tip": "Highlight your market analysis of Kuvera and Dezerv from your internship."
        },
        {
            "title": "Business Analyst - Data & AI",
            "company": "Valtech (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/valtech",
            "fits": "High-growth digital agency. Leverages data visualization skills.",
            "tip": "Emphasize Tata Data Visualisation certification and dashboard understanding."
        },
        {
            "title": "Lead Business Intelligence Analyst",
            "company": "Cision (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/cision",
            "fits": "Modern BI stack. Fits your statistical analysis projects.",
            "tip": "Highlight correlation of working capital with profitability in FMCG study."
        },
        {
            "title": "Data QA Associate",
            "company": "YipitData (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://job-boards.greenhouse.io/yipitdatajobs",
            "fits": "Data cleaning and quality assurance. Matches Excel/data skills.",
            "tip": "Highlight database operations, Excel, and data-driven projects."
        },
        {
            "title": "Junior Business Analyst",
            "company": "AlphaSense (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/alphasense",
            "fits": "Financial search engine support. Perfect for B.Com graduates.",
            "tip": "Showcase your competitive benchmarking of Dezerv and Kuvera."
        },
        {
            "title": "Business Research Intern",
            "company": "Perccent (Return Opp)",
            "category": "Business Analysis",
            "employment": "Part-time",
            "url": "https://manishbhandari.framer.website/",
            "fits": "Check for return/contract extensions at Perccent.",
            "tip": "Leverage your existing relationships and success at the firm."
        },
        {
            "title": "Business Analyst Intern",
            "company": "Unstop Jobs (India)",
            "category": "Business Analysis",
            "employment": "Part-time",
            "url": "https://unstop.com/internships?specialisation=Business%20Analysis",
            "fits": "Fresher internships in business analysis. Part-time.",
            "tip": "Highlight Excel, SQL, and Notion proficiency in your skills section."
        },
        {
            "title": "Business Analyst Remote",
            "company": "Instahyre (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://www.instahyre.com/jobs-in-india/?search=Business+Analyst",
            "fits": "Remote startup business analyst listings in India.",
            "tip": "Highlight your corporate accounting and financial system courses."
        },
        {
            "title": "Junior Analyst",
            "company": "Cutshort (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://cutshort.io/jobs/remote-business-analyst-jobs",
            "fits": "Remote BA roles on Cutshort.",
            "tip": "Focus on your analytical skills and Perccent experience."
        },
        {
            "title": "Data Analyst (Remote)",
            "company": "Indeed India",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://in.indeed.com/q-data-analyst-l-remote-jobs.html",
            "fits": "Aggregated data roles in India.",
            "tip": "Emphasize data visualization and working capital statistical study."
        },
        {
            "title": "Business Intelligence Intern",
            "company": "Shadowfax (India)",
            "category": "Business Analysis",
            "employment": "Part-time",
            "url": "https://www.shadowfax.in/careers",
            "fits": "Logistics tech company. Part-time/flexible analytics internship.",
            "tip": "Showcase your data visualization skills and working capital FMCG study."
        },
        {
            "title": "Business Analyst",
            "company": "InVideo (India)",
            "category": "Business Analysis",
            "employment": "Full-time",
            "url": "https://invideo.io/careers/",
            "fits": "Remote-first video creation SaaS. High growth startup.",
            "tip": "Showcase your ability to write detailed user stories and translate specs."
        },

        # --- FINANCE & OPERATIONS (12 entries) ---
        {
            "title": "Operations Associate",
            "company": "Kuvera (CRED) (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://grow.kuvera.in/",
            "fits": "Wealth operations. Fits your M&A Strategic Analysis of CRED x Kuvera.",
            "tip": "Emphasize your understanding of NetBanking and UPI mandate flows."
        },
        {
            "title": "Associate Portfolio Data Analyst",
            "company": "Addepar (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/addepar",
            "fits": "Wealth management technology. Perfectly matches B.Com and WealthTech.",
            "tip": "Highlight corporate accounting courses and Perccent internship."
        },
        {
            "title": "Alts Data Operations Analyst",
            "company": "Addepar (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/addepar",
            "fits": "Alternative investments. Aligns with minor/NRI investor research.",
            "tip": "Showcase your research on fee models and untapped investor segments."
        },
        {
            "title": "Procurement Analyst",
            "company": "Precision Medicine Group (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/precisionmedicinegroup",
            "fits": "Finance and purchasing. Fits B.Com degree.",
            "tip": "Emphasize corporate accounting and tax courses."
        },
        {
            "title": "Graduate Financial Analyst",
            "company": "Zomato (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://www.zomato.com/careers",
            "fits": "Entry-level finance. Fits B.Com education.",
            "tip": "Highlight your Working Capital Management analysis project."
        },
        {
            "title": "Finance Operations Associate",
            "company": "Razorpay (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://razorpay.com/jobs/",
            "fits": "Payment gateway operations. Fits netbanking/UPI spec project.",
            "tip": "Showcase your spec writing for UPI mandate and netbanking flows."
        },
        {
            "title": "Operations Associate",
            "company": "Paytm (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://careers.paytm.com/",
            "fits": "Wealth and fintech operations. Perfect for B.Com.",
            "tip": "Highlight your Perccent research on AUM trends and fee models."
        },
        {
            "title": "Junior Finance Analyst",
            "company": "Paytm Money (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://careers.paytm.com/",
            "fits": "Mutual funds and wealth management. Fits B.Com.",
            "tip": "Emphasize your Step-Up SIP and CAGR toggle acceptance criteria."
        },
        {
            "title": "Finance Intern",
            "company": "Unstop Jobs (India)",
            "category": "Finance & Operations",
            "employment": "Part-time",
            "url": "https://unstop.com/internships?specialisation=Finance",
            "fits": "Remote finance internships for B.Com students. Part-time.",
            "tip": "Highlight SGPA: 8.10 and corporate accounting academic work."
        },
        {
            "title": "Financial Analyst Remote",
            "company": "Indeed India",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://in.indeed.com/q-financial-analyst-l-remote-jobs.html",
            "fits": "Aggregated remote finance roles in India.",
            "tip": "Highlight FMCG Working Capital project (HUL, ITC, Nestle)."
        },
        {
            "title": "Finance Intern",
            "company": "Zepto (India)",
            "category": "Finance & Operations",
            "employment": "Part-time",
            "url": "https://www.zepto.co.in/careers",
            "fits": "Quick commerce. Part-time finance/working capital internship.",
            "tip": "Highlight your FMCG Working Capital project (CCC and turnover cycles)."
        },
        {
            "title": "Operations Analyst",
            "company": "InVideo (India)",
            "category": "Finance & Operations",
            "employment": "Full-time",
            "url": "https://invideo.io/careers/",
            "fits": "Operations management. Fits outreach and Notion skills.",
            "tip": "Highlight structured problem-solving and client outreach skills."
        },

        # --- MARKETING, CONTENT & GROWTH (10 entries) ---
        {
            "title": "Growth Marketing Intern",
            "company": "Classplus (India)",
            "category": "Marketing & Growth",
            "employment": "Part-time",
            "url": "https://classplusapp.com/careers",
            "fits": "EdTech growth and marketing. Part-time/flexible hours.",
            "tip": "Highlight Canva, Framer, and LinkedIn content creation."
        },
        {
            "title": "Growth Associate (Intern)",
            "company": "Unstop Jobs (India)",
            "category": "Marketing & Growth",
            "employment": "Part-time",
            "url": "https://unstop.com/internships?specialisation=Marketing",
            "fits": "Remote marketing/growth internships in India. Part-time.",
            "tip": "Showcase Google Fundamentals of Marketing certification."
        },
        {
            "title": "Growth Marketing Associate",
            "company": "Simpl (India)",
            "category": "Marketing & Growth",
            "employment": "Full-time",
            "url": "https://getsimpl.com/careers",
            "fits": "Fintech growth and user acquisition.",
            "tip": "Highlight social media marketing content created for Perccent."
        },
        {
            "title": "Marketing Analyst",
            "company": "YipitData (India)",
            "category": "Marketing & Growth",
            "employment": "Full-time",
            "url": "https://job-boards.greenhouse.io/yipitdatajobs",
            "fits": "Marketing data analysis. Fits marketing certification & data.",
            "tip": "Emphasize Tata Data Visualisation and Google Marketing."
        },
        {
            "title": "Content Marketing Associate",
            "company": "Groww (India)",
            "category": "Marketing & Growth",
            "employment": "Full-time",
            "url": "https://groww.in/careers",
            "fits": "WealthTech education and content. Fits creative writing.",
            "tip": "Mention LinkedIn and Instagram content management at Perccent."
        },
        {
            "title": "Growth Associate",
            "company": "Paytm (India)",
            "category": "Marketing & Growth",
            "employment": "Full-time",
            "url": "https://careers.paytm.com/",
            "fits": "Fintech marketing and growth. Fits marketing background.",
            "tip": "Mention Google Fundamentals of Marketing and Canva/Framer skills."
        },
        {
            "title": "Growth Associate",
            "company": "Instahyre (India)",
            "category": "Marketing & Growth",
            "employment": "Full-time",
            "url": "https://www.instahyre.com/jobs-in-india/?search=Marketing+Associate",
            "fits": "Remote startup marketing roles on Instahyre.",
            "tip": "Emphasize brand engagement metrics driven during Perccent internship."
        },
        {
            "title": "Digital Marketing Intern",
            "company": "Internshala (India)",
            "category": "Marketing & Growth",
            "employment": "Part-time",
            "url": "https://internshala.com/internships/work-from-home-marketing-internships/",
            "fits": "Remote marketing internships in India. Part-time.",
            "tip": "Show your portfolio website and creative writing samples."
        },
        {
            "title": "Social Media & Design Intern",
            "company": "InVideo (India)",
            "category": "Marketing & Growth",
            "employment": "Part-time",
            "url": "https://invideo.io/careers/",
            "fits": "Social media content and Framer/Canva design. Part-time.",
            "tip": "Pitch your Perccent brand engagement metrics and Framer website link."
        },
        {
            "title": "Marketing Associate",
            "company": "Classplus (India)",
            "category": "Marketing & Growth",
            "employment": "Full-time",
            "url": "https://classplusapp.com/careers",
            "fits": "Edtech platform marketing. Fits outreach & writing.",
            "tip": "Highlight outreach and client communication soft skills."
        },

        # --- BUSINESS DEVELOPMENT, OUTREACH & SALES (10 entries) ---
        {
            "title": "Business Development Associate",
            "company": "upGrad (India)",
            "category": "Sales & BD",
            "employment": "Full-time",
            "url": "https://www.upgrad.com/careers/",
            "fits": "Higher education sales. Fits PepsiCo Sales simulation.",
            "tip": "Highlight your PepsiCo Sales Job Simulation certification."
        },
        {
            "title": "Sales Operations Associate",
            "company": "Razorpay (India)",
            "category": "Sales & BD",
            "employment": "Full-time",
            "url": "https://razorpay.com/jobs/",
            "fits": "Fintech sales support and client outreach.",
            "tip": "Highlight client communication and outreach soft skills."
        },
        {
            "title": "Business Development Intern",
            "company": "Unstop Jobs (India)",
            "category": "Sales & BD",
            "employment": "Part-time",
            "url": "https://unstop.com/internships?specialisation=Business%20Development",
            "fits": "BD internships with remote options. Part-time.",
            "tip": "Emphasize leadership and client communication."
        },
        {
            "title": "Client Success Associate",
            "company": "AlphaSense (India)",
            "category": "Sales & BD",
            "employment": "Full-time",
            "url": "https://boards.greenhouse.io/alphasense",
            "fits": "Financial platform customer onboarding. Fits B.Com.",
            "tip": "Show how your Perccent experience helps you understand client needs."
        },
        {
            "title": "Inside Sales Associate",
            "company": "upGrad (India)",
            "category": "Sales & BD",
            "employment": "Full-time",
            "url": "https://www.upgrad.com/careers/",
            "fits": "Outbound student counseling and sales.",
            "tip": "Highlight outreach and leadership soft skills."
        },
        {
            "title": "Sales Intern",
            "company": "PepsiCo (India) (Return)",
            "category": "Sales & BD",
            "employment": "Part-time",
            "url": "https://www.pepsicojobs.com/",
            "fits": "Keep an eye on PepsiCo careers. Fits Sales simulation. Part-time.",
            "tip": "Lead with your PepsiCo Sales Job Simulation completion."
        },
        {
            "title": "Business Development Associate",
            "company": "Instahyre (India)",
            "category": "Sales & BD",
            "employment": "Full-time",
            "url": "https://www.instahyre.com/jobs-in-india/?search=Business+Development+Associate",
            "fits": "Startup sales roles on Instahyre in India.",
            "tip": "Emphasize outreach and problem-solving skills."
        },
        {
            "title": "Business Development Intern",
            "company": "Internshala (India)",
            "category": "Sales & BD",
            "employment": "Part-time",
            "url": "https://internshala.com/internships/work-from-home-business-development-internships/",
            "fits": "Remote BD internships for students in India. Part-time.",
            "tip": "Highlight sales simulation and client communication skills."
        },
        {
            "title": "Customer Success Analyst",
            "company": "Pocket FM (India)",
            "category": "Sales & BD",
            "employment": "Full-time",
            "url": "https://www.pocketfm.com/careers",
            "fits": "Remote customer success roles at audio platform.",
            "tip": "Highlight problem-solving and creative writing."
        },
        {
            "title": "Business Development Associate",
            "company": "Indeed India",
            "category": "Sales & BD",
            "employment": "Full-time",
            "url": "https://in.indeed.com/q-business-development-associate-l-remote-jobs.html",
            "fits": "Aggregated BD roles in India.",
            "tip": "Emphasize communication skills and sales training."
        }
    ]
    
    start_row = 5
    for idx, job in enumerate(jobs):
        curr_row = start_row + idx
        ws.row_dimensions[curr_row].height = 35
        
        # Alternating fills (Zebra striping)
        row_fill = fill_zebra if idx % 2 == 1 else fill_white
        
        # 1. Job Title
        cell_title = ws.cell(row=curr_row, column=1, value=job["title"])
        cell_title.font = Font(name="Segoe UI", size=10, bold=True)
        cell_title.alignment = align_left
        
        # 2. Company / Platform
        cell_comp = ws.cell(row=curr_row, column=2, value=job["company"])
        cell_comp.font = font_body
        cell_comp.alignment = align_left
        
        # 3. Role Category
        cell_cat = ws.cell(row=curr_row, column=3, value=job["category"])
        cell_cat.font = font_body
        cell_cat.alignment = align_center
        
        # 4. Employment Type
        cell_emp = ws.cell(row=curr_row, column=4, value=job["employment"])
        cell_emp.font = Font(name="Segoe UI", size=10, bold=(job["employment"] == "Part-time"))
        cell_emp.alignment = align_center
        
        # 5. Application Link (Clickable hyperlink)
        cell_link = ws.cell(row=curr_row, column=5, value="Apply Link ↗")
        cell_link.hyperlink = job["url"]
        cell_link.font = font_link
        cell_link.alignment = align_center
        
        # 6. Activity Status (Default is 'Pending')
        cell_status = ws.cell(row=curr_row, column=6, value="Pending")
        cell_status.font = Font(name="Segoe UI", size=10, bold=True)
        cell_status.alignment = align_center
        
        # 7. Date Discovered
        cell_date = ws.cell(row=curr_row, column=7, value="2026-06-08")
        cell_date.font = font_body
        cell_date.alignment = align_center
        
        # 8. Why it Fits
        cell_fits = ws.cell(row=curr_row, column=8, value=job["fits"])
        cell_fits.font = font_body
        cell_fits.alignment = align_left
        
        # 9. Strategy & Tip
        cell_tip = ws.cell(row=curr_row, column=9, value=job["tip"])
        cell_tip.font = font_body
        cell_tip.alignment = align_left
        
        # Apply borders and backgrounds to all cells in the row
        for col_idx in range(1, 10):
            c = ws.cell(row=curr_row, column=col_idx)
            c.border = border_all
            if col_idx not in (4, 6):  # Let status and employment stand out without zebra/white defaults to be highly readable
                c.fill = row_fill
            elif col_idx == 4:
                # Add light blue fill for employment column
                c.fill = PatternFill(start_color="F0F4C3" if job["employment"] == "Part-time" else "E8F5E9", end_color="F0F4C3" if job["employment"] == "Part-time" else "E8F5E9", fill_type="solid")
            else:
                # Add light yellow fill for status column
                c.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                
    # 2. Add Dropdown Data Validation for Status Column (Col 6, Rows 5 to 60)
    dv_status = DataValidation(
        type="list", 
        formula1='"Pending,Applied,Interviewing,Offered,Rejected"', 
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    for r in range(start_row, start_row + len(jobs)):
        dv_status.add(ws.cell(row=r, column=6))
        
    # 3. Add Dropdown Data Validation for Employment Type Column (Col 4, Rows 5 to 60)
    dv_emp = DataValidation(
        type="list", 
        formula1='"Full-time,Part-time"', 
        allow_blank=True
    )
    ws.add_data_validation(dv_emp)
    for r in range(start_row, start_row + len(jobs)):
        dv_emp.add(ws.cell(row=r, column=4))
        
    # Auto-fit columns with safety margins
    column_widths = {
        "A": 24,  # Job Title
        "B": 24,  # Company / Platform (India)
        "C": 20,  # Role Category
        "D": 18,  # Employment Type
        "E": 14,  # Application Link
        "F": 16,  # Activity Status
        "G": 16,  # Date Discovered
        "H": 38,  # Why it Fits Your Resume
        "I": 38   # Strategy & Tip
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Save output
    output_filename = "job_tracker.xlsx"
    wb.save(output_filename)
    print(f"Spreadsheet generated successfully with Dashboard and Tracker tabs: {output_filename}")

if __name__ == "__main__":
    create_tracker()
